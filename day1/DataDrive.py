import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
import requests
import time
from datetime import datetime, timedelta
from selenium.webdriver import Keys, ActionChains
from selenium.webdriver.support.select import Select
from openpyxl import load_workbook
# for login function we are going to pass it through different users

#creating the excel sheet after that load workbook

workbook =  load_workbook("automation.xlsx")
sheet = workbook.active
driver = webdriver.Chrome()
driver.maximize_window()
# driver.get("https://www.saucedemo.com/v1/")

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]

    driver.get("https://www.saucedemo.com/v1/")
    time.sleep(3)
    driver.find_element(By.ID,"user-name").send_keys(username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"login-button").click()
    time.sleep(6)
    driver.find_element(By.XPATH,"//button[normalize-space()='Open Menu']").click()
    time.sleep(6)
    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
    time.sleep(6)
driver.quit()